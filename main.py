from selenium import webdriver
from selenium.webdriver import ActionChains, FirefoxProfile
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC, expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChain

wb = load_workbook(r"C:\users\leoce\Desktop\Mappe1.xlsx")
sheet_obj = wb.active
m_row = sheet_obj.max_row
firefox_options = webdriver.FirefoxOptions()
browser = webdriver.Firefox(options=firefox_options)
email1_obj = sheet_obj.cell(row=2, column=1)
pw1_obj = sheet_obj.cell(row=2, column=2)

browser.get(
    "https://www.amazon.de/ring-video-doorbell-wired/dp/B08LR3G17D/ref=sr_1_2?keywords=ring&qid=1652445690&sr=8-2")

cookieButton = browser.find_element(By.XPATH, '//*[@id="sp-cc-accept"]')
buyButton = browser.find_element(By.XPATH, '//*[@id="buy-now-button"]')

cookieButton.click()
buyButton.click()

emailField = browser.find_element(By.XPATH, '//*[@id="ap_email"]')
emailField.send_keys(email1_obj.value + Keys.ENTER)

browser.implicitly_wait(2)

passwordField = browser.find_element(By.CSS_SELECTOR, '#ap_password')
passwordField.send_keys(pw1_obj.value + Keys.ENTER)

for i in range(2, m_row + 1):
    browser.switch_to.new_window('tab')
    browser.get(
        "https://www.amazon.de/ring-video-doorbell-wired/dp/B08LR3G17D/ref=sr_1_2?keywords=ring&qid=1652445690&sr=8-2")

    buyButton2 = browser.find_element(By.XPATH, '//*[@id="buy-now-button"]')
    buyButton2.click()
    browser.implicitly_wait(50)
    WebDriverWait(browser, 20).until(expected_conditions.element_to_be_clickable((By.XPATH,
                                                        '//*[@id="address-ui-widgets-enterAddressPhoneNumber"]')))
    browser.back()

    logoutHover = browser.find_element(By.XPATH, '//*[@id="nav-link-accountList"]')
    hover = ActionChains(browser).move_to_element(logoutHover)
    hover.perform()
    WebDriverWait(browser, 20).until(expected_conditions.element_to_be_clickable((By.XPATH,
                                                                            '//*[@id="nav-item-signout"]'))).click()


    emailField = browser.find_element(By.XPATH, '//*[@id="ap_email"]')
    emailObject = sheet_obj.cell(row=i + 1, column=1)
    emailField.send_keys(emailObject.value + Keys.ENTER)
